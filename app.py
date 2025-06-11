from flask import Flask, request, render_template, jsonify, Response
import pandas as pd
import numpy as np
import joblib
import os
import sqlite3
from sklearn.preprocessing import LabelEncoder # Pastikan ini di-import!
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

app = Flask(__name__)

# --- Load the trained model ---
try:
    # IMPORTANT: Ensure 'best_tuned_model_kpr_fix.pkl' is the correct model file
    # If your original model was 'soft_voting_model.pkl' from the other project,
    # and this project uses a different model/dataset, confirm the filename.
    model = joblib.load('best_tuned_model_kpr_fix.pkl') 
except FileNotFoundError:
    print("Error: 'best_tuned_model_kpr_fix.pkl' not found. Make sure the model file is in the same directory.")
    exit()

# --- Replikasi LabelEncoders dari training di ML_KPR_3.ipynb ---
# These are for your original KPR project, not the "other project" (Home Credit)
label_encoders_replication = {}
categorical_cols_for_le = ['Gender', 'Married', 'Education', 'Self_Employed', 'Property_Area']

LE_CLASSES = {
    'Gender': ['Female', 'Male'],
    'Married': ['No', 'Yes'],
    'Education': ['Graduate', 'Not Graduate'],
    'Self_Employed': ['No', 'Yes'],
    'Property_Area': ['Rural', 'Semiurban', 'Urban']
}

for col in categorical_cols_for_le:
    le = LabelEncoder()
    le.fit(LE_CLASSES[col])
    label_encoders_replication[col] = le

# --- Manual Label Encoding Dictionaries (from ML_KPR_3.ipynb) ---
# These are for your original KPR project, not the "other project" (Home Credit)
label_maps_kpr_project = { # Renamed to avoid conflict with other project's label_maps
    'Gender': {
        'Male': 1,
        'Female': 0
    },
    'Married': {
        'Yes': 1,
        'No': 0
    },
    'Dependents': { 
        '0': 0,
        '1': 1,
        '2': 2,
        '3+': 3 
    },
    'Education': {
        'Graduate': 1,
        'Not Graduate': 0
    },
    'Self_Employed': {
        'Yes': 1,
        'No': 0
    },
    'Property_Area': {
        'Urban': 2,
        'Semiurban': 1,
        'Rural': 0
    },
    'Credit_History': { 
        '0': 0,
        '1': 1
    }
}

# --- Mapping manual Loan_Amount_Term to Loan_Amount_Term_Code ---
term_mapping = {
    12.0: 1,
    36.0: 2,
    60.0: 3,
    84.0: 4,
    120.0: 5,
    180.0: 6,
    240.0: 7,
    300.0: 8,
    360.0: 9,
    480.0: 10
}

# --- Default Imputation Values (from ML_KPR_3.ipynb, based on mode/median) ---
# Updated default values to be in IDR instead of USD
DEFAULT_IMPUTATION_VALUES = {
    'Gender': 'Male',
    'Married': 'Yes',
    'Dependents': '0',
    'Education': 'Graduate',
    'Self_Employed': 'No',
    'Credit_History': '1', 
    'ApplicantIncome': 5000000.0,  # 5 juta IDR per bulan
    'CoapplicantIncome': 0.0, 
    'LoanAmount': 150000000.0,  # 150 juta IDR (bukan dalam ribuan)
    'Loan_Amount_Term': 360.0, 
}

# --- KPR Simulation Parameters ---
KPR_ANNUAL_INTEREST_RATE = 7.5 # in percent (e.g., 7.5% per year)

# --- Fungsi untuk koneksi database SQLite ---
def get_db_connection():
    conn = sqlite3.connect('house.db')
    conn.row_factory = sqlite3.Row
    return conn

# --- Fungsi untuk menghitung simulasi KPR (Annuitas) ---
def calculate_kpr_simulation_details(principal_loan_idr, loan_term_months, annual_interest_rate_percent):
    if principal_loan_idr <= 0 or loan_term_months <= 0 or annual_interest_rate_percent < 0:
        return {
            "monthly_payment": 0.0,
            "total_interest_paid": 0.0,
            "total_payment": 0.0,
            "annual_interest_rate": annual_interest_rate_percent # Still return the rate used
        }

    monthly_interest_rate = (annual_interest_rate_percent / 100) / 12
    
    # Annuity formula for fixed-rate loans
    if monthly_interest_rate == 0: # Handle 0% interest case
        monthly_payment = principal_loan_idr / loan_term_months
    else:
        # M = P [ i(1 + i)^n ] / [ (1 + i)^n – 1]
        monthly_payment = principal_loan_idr * (monthly_interest_rate * (1 + monthly_interest_rate)**loan_term_months) / ((1 + monthly_interest_rate)**loan_term_months - 1)
    
    total_payment = monthly_payment * loan_term_months
    total_interest_paid = total_payment - principal_loan_idr

    return {
        "monthly_payment": monthly_payment,
        "total_interest_paid": total_interest_paid,
        "total_payment": total_payment,
        "annual_interest_rate": annual_interest_rate_percent
    }

# --- Fungsi untuk menghitung detail cicilan per bulan ---
def calculate_monthly_installments(principal_loan_idr, loan_term_months, annual_interest_rate_percent):
    monthly_interest_rate = (annual_interest_rate_percent / 100) / 12
    monthly_payment = principal_loan_idr * (monthly_interest_rate * (1 + monthly_interest_rate)**loan_term_months) / ((1 + monthly_interest_rate)**loan_term_months - 1)
    
    installments = []
    remaining_principal = principal_loan_idr
    
    for month in range(1, loan_term_months + 1):
        interest_payment = remaining_principal * monthly_interest_rate
        principal_payment = monthly_payment - interest_payment
        remaining_principal -= principal_payment
        
        installments.append({
            'Bulan_ke': month,
            'Cicilan_Bulanan': monthly_payment,
            'Pembayaran_Pokok': principal_payment,
            'Pembayaran_Bunga': interest_payment,
            'Sisa_Pinjaman': max(0, remaining_principal)  # Ensure we don't show negative values
        })
    
    return installments

# --- Pra-pemrosesan input untuk MODEL KPR (menggunakan data dari properties-single.html) ---
def preprocess_kpr_input(form_data):
    # This dictionary will hold the data specifically for the model's input features.
    processed_data_for_model = {} 
    
    # This dictionary will hold the original or calculated IDR values for display on hasil-simulasi.html
    display_data_for_frontend = {}

    # --- Helper untuk mendapatkan nilai numerik dari form (Rupiah robust parsing) ---
    def get_raw_number_from_formatted_idr(s):
        if isinstance(s, (int, float)): return s
        try: 
            s_str = str(s).strip()
            if not s_str: # Handle empty string explicitly
                return np.nan
            
            # Safest approach for large Rupiah numbers: remove all non-digits
            cleaned_s = ''.join(filter(str.isdigit, s_str))
            
            if not cleaned_s:
                return np.nan

            return float(cleaned_s)
        except ValueError: 
            return np.nan # Return NaN if not convertible to float

    # --- 1. Ambil & Proses Nilai Numerik dari Form (dalam Rupiah) ---
    raw_applicant_income_idr = get_raw_number_from_formatted_idr(form_data.get('ApplicantIncome', None))
    raw_coapplicant_income_idr = get_raw_number_from_formatted_idr(form_data.get('CoapplicantIncome', None))
    raw_harga_rumah_idr = get_raw_number_from_formatted_idr(form_data.get('harga_rumah', None))
    raw_uang_muka_idr = get_raw_number_from_formatted_idr(form_data.get('uang_muka', None))
    raw_loan_amount_term_months = float(form_data.get('Loan_Amount_Term', 0))  # Get as float

    # Store loan term data for display
    display_data_for_frontend['Loan_Amount_Term'] = int(raw_loan_amount_term_months)
    display_data_for_frontend['Loan_Amount_Term_Code'] = form_data.get('Loan_Amount_Term_Code', str(int(raw_loan_amount_term_months)))

    # Store other display data
    display_data_for_frontend['HargaRumah'] = raw_harga_rumah_idr
    display_data_for_frontend['UangMuka'] = raw_uang_muka_idr
    display_data_for_frontend['JumlahPinjamanDiajukan'] = raw_harga_rumah_idr - raw_uang_muka_idr
    display_data_for_frontend['ApplicantIncome'] = raw_applicant_income_idr
    display_data_for_frontend['CoapplicantIncome'] = raw_coapplicant_income_idr

    # --- Imputasi & Tentukan Nilai Akhir untuk Display dan Model Calculation ---

    # ApplicantIncome (IDR for both display and model)
    applicant_income_idr_val = raw_applicant_income_idr if not pd.isna(raw_applicant_income_idr) else DEFAULT_IMPUTATION_VALUES['ApplicantIncome']
    display_data_for_frontend['ApplicantIncome'] = applicant_income_idr_val 
    processed_data_for_model['ApplicantIncome'] = applicant_income_idr_val

    # CoapplicantIncome (IDR for both display and model)
    coapplicant_income_idr_val = raw_coapplicant_income_idr if not pd.isna(raw_coapplicant_income_idr) else DEFAULT_IMPUTATION_VALUES['CoapplicantIncome']
    display_data_for_frontend['CoapplicantIncome'] = coapplicant_income_idr_val 
    processed_data_for_model['CoapplicantIncome'] = coapplicant_income_idr_val

    # Harga Rumah & Uang Muka (IDR for display)
    harga_rumah_idr_val = raw_harga_rumah_idr if not pd.isna(raw_harga_rumah_idr) and raw_harga_rumah_idr > 0 else 0.0
    uang_muka_idr_val = raw_uang_muka_idr if not pd.isna(raw_uang_muka_idr) and raw_uang_muka_idr >= 0 else 0.0

    display_data_for_frontend['HargaRumah'] = harga_rumah_idr_val 
    display_data_for_frontend['UangMuka'] = uang_muka_idr_val 

    # Pokok Pinjaman Diajukan (IDR for display & KPR calculation)
    pokok_pinjaman_diajukan_idr = harga_rumah_idr_val - uang_muka_idr_val
    if pokok_pinjaman_diajukan_idr <= 0: # Ensure loan amount is positive
        pokok_pinjaman_diajukan_idr = DEFAULT_IMPUTATION_VALUES['LoanAmount']
    display_data_for_frontend['JumlahPinjamanDiajukan'] = pokok_pinjaman_diajukan_idr 

    # LoanAmount for model (in IDR, no conversion to USD thousands)
    processed_data_for_model['LoanAmount'] = pokok_pinjaman_diajukan_idr

    # Loan_Amount_Term (numerik, dalam bulan, for display & model)
    loan_amount_term_months_val = raw_loan_amount_term_months if not pd.isna(raw_loan_amount_term_months) and raw_loan_amount_term_months > 0 else DEFAULT_IMPUTATION_VALUES['Loan_Amount_Term']
    
    display_data_for_frontend['Loan_Amount_Term'] = loan_amount_term_months_val 
    processed_data_for_model['Loan_Amount_Term'] = loan_amount_term_months_val 

    # Add Loan_Amount_Term_Code based on term_mapping (for model)
    processed_data_for_model['Loan_Amount_Term_Code'] = term_mapping.get(float(loan_amount_term_months_val), 0) 


    # --- Apply Log Transformation to numerical features for the model ---
    # It's crucial to apply log transformation *after* imputation and ensuring non-negative values.
    # Use np.log1p (log(1+x)) to handle potential zero values gracefully.
    processed_data_for_model['ApplicantIncome'] = np.log1p(processed_data_for_model['ApplicantIncome'])
    processed_data_for_model['CoapplicantIncome'] = np.log1p(processed_data_for_model['CoapplicantIncome'])
    processed_data_for_model['LoanAmount'] = np.log1p(processed_data_for_model['LoanAmount'])


    # --- 2. Proses Fitur Kategorikal dan Imputasi + Encoding ---
    categorical_features = ['Gender', 'Married', 'Dependents', 'Education', 'Self_Employed', 'Credit_History', 'Property_Area']
    for feature in categorical_features:
        value = form_data.get(feature)
        
        # Imputasi nilai hilang dengan default string/numerik sesuai tipe
        if value is None or (isinstance(value, str) and value.strip() == ''):
            imputed_val = DEFAULT_IMPUTATION_VALUES.get(feature, list(label_maps_kpr_project[feature].keys())[0])
            processed_data_for_model[feature] = imputed_val 
            display_data_for_frontend[feature] = imputed_val 
        else:
            processed_data_for_model[feature] = value 
            display_data_for_frontend[feature] = value 

        # Terapkan Encoding sesuai strategi di notebook
        if feature == 'Dependents': 
            dep_map = label_maps_kpr_project['Dependents']
            if isinstance(processed_data_for_model[feature], (int, float)): 
                processed_data_for_model[feature] = int(processed_data_for_model[feature])
            else: 
                processed_data_for_model[feature] = dep_map.get(processed_data_for_model[feature], dep_map['0']) 
        
        elif feature == 'Credit_History': 
            if processed_data_for_model[feature] is not None and processed_data_for_model[feature] != '':
                try:
                    processed_data_for_model[feature] = int(float(processed_data_for_model[feature]))
                except ValueError: 
                    processed_data_for_model[feature] = int(DEFAULT_IMPUTATION_VALUES['Credit_History'])
            else: 
                processed_data_for_model[feature] = int(DEFAULT_IMPUTATION_VALUES['Credit_History'])
        
        else: # Gunakan LabelEncoder untuk fitur-fitur kategorikal lainnya
            try:
                processed_data_for_model[feature] = label_encoders_replication[feature].transform([processed_data_for_model[feature]])[0]
            except ValueError:
                print(f"Warning: Category '{processed_data_for_model[feature]}' for feature '{feature}' not seen by encoder. Defaulting to encoded imputation default.")
                processed_data_for_model[feature] = label_encoders_replication[feature].transform([DEFAULT_IMPUTATION_VALUES[feature]])[0]


    # --- 3. Feature Engineering (Sesuai ML_KPR_3.ipynb Anda) ---
    # These calculations should use the log-transformed values if they were used for training
    # For example, if Total_Income in training was sum of log(ApplicantIncome) + log(CoapplicantIncome),
    # then apply it here. If it was sum of original incomes, then use original incomes for this.
    # Assuming Total_Income was calculated from log-transformed ApplicantIncome and CoapplicantIncome for the model.
    processed_data_for_model['Total_Income'] = processed_data_for_model['ApplicantIncome'] + processed_data_for_model['CoapplicantIncome']
    
    # Ensure Dependents is treated as a number for division
    dependents_numeric = processed_data_for_model['Dependents']
    if isinstance(dependents_numeric, str) and dependents_numeric == '3+':
        dependents_numeric = 3
    else:
        try:
            dependents_numeric = int(dependents_numeric)
        except ValueError:
            dependents_numeric = int(label_maps_kpr_project['Dependents']['0']) # Use default '0' mapped value

    if (dependents_numeric + 1) == 0:
        processed_data_for_model['Income_Per_Person'] = processed_data_for_model['Total_Income']
    else:
        processed_data_for_model['Income_Per_Person'] = processed_data_for_model['Total_Income'] / (dependents_numeric + 1)

    if processed_data_for_model['Total_Income'] == 0:
        processed_data_for_model['Loan_Income_Ratio'] = 0.0
    else:
        # If LoanAmount was log-transformed for Loan_Income_Ratio calculation during training, use its log-transformed value
        # Otherwise, use the original (but imputed) LoanAmount for this ratio if that's how it was trained.
        # Based on typical usage, if LoanAmount is log-transformed, Loan_Income_Ratio would also use log-transformed values.
        processed_data_for_model['Loan_Income_Ratio'] = processed_data_for_model['LoanAmount'] / processed_data_for_model['Total_Income']

    processed_data_for_model['Educated_SelfEmployed'] = processed_data_for_model['Education'] * processed_data_for_model['Self_Employed']
    # Ensure 'Married' and 'Dependents' are using their encoded integer values for comparison
    processed_data_for_model['Is_Single'] = int((processed_data_for_model['Married'] == label_maps_kpr_project['Married']['No']) and (processed_data_for_model['Dependents'] == label_maps_kpr_project['Dependents']['0']))

    # 4. Buat DataFrame final untuk model
    final_input_df = pd.DataFrame([processed_data_for_model])

    # --- FINAL MODEL FEATURES dari ML_KPR_3.ipynb Anda ---
    MODEL_FEATURES = [
        'Gender', 'Married', 'Dependents', 'Education', 'Self_Employed',
        'ApplicantIncome', 'CoapplicantIncome', 'LoanAmount', 'Credit_History',
        'Property_Area', 'Loan_Amount_Term_Code', 'Total_Income',
        'Income_Per_Person', 'Loan_Income_Ratio', 'Educated_SelfEmployed', 'Is_Single'
    ]

    # Pastikan DataFrame memiliki semua MODEL_FEATURES dalam urutan yang benar
    final_input_df = final_input_df[MODEL_FEATURES]

    return {
        "model_input": final_input_df,
        "display_data": display_data_for_frontend # Mengembalikan detail pinjaman untuk tampilan
    }


# --- Fungsi untuk mendapatkan rekomendasi rumah dari database ---
def get_recommended_houses(user_income_total_idr, pokok_pinjaman_idr, property_area_pref=None): # Removed tenor_years, required_bedrooms, required_bathrooms as they are not used in new KPR project context
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM houses")
    all_houses = [dict(row) for row in cursor.fetchall()]
    conn.close()

    recommended_list = []
    
    # Aturan Rekomendasi Sederhana:
    # Based on loan amount and income
    max_house_price_loan = pokok_pinjaman_idr * 1.5
    min_house_price_loan = pokok_pinjaman_idr * 0.8 

    # Assuming user_income_total_idr is combined applicant + coapplicant income IDR
    # Adjust multipliers as needed for better recommendations
    max_house_price_income = user_income_total_idr * 5 
    min_house_price_income = user_income_total_idr * 3 

    for house in all_houses:
        price_fits_loan = (min_house_price_loan <= house['harga_idr'] <= max_house_price_loan)
        price_fits_income = (min_house_price_income <= house['harga_idr'] <= max_house_price_income)
        
        if price_fits_loan and price_fits_income:
            location_match = True 
            if property_area_pref: 
                house_location_lower = house['lokasi'].lower()
                if property_area_pref == 'Urban' and not ('kota' in house_location_lower or 'urban' in house_location_lower or 'city' in house_location_lower):
                    location_match = False
                elif property_area_pref == 'Semiurban' and not ('semi' in house_location_lower or 'suburban' in house_location_lower):
                    location_match = False
                elif property_area_pref == 'Rural' and not ('desa' in house_location_lower or 'rural' in house_location_lower):
                    location_match = False
            
            if location_match:
                recommended_list.append(house)
    
    recommended_list = sorted(recommended_list, key=lambda x: x['harga_idr'])
    
    return recommended_list[:6] # Mengambil hingga 6 rekomendasi teratas


# --- Flask Routes ---
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/properties')
def properties_page():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM houses ORDER BY harga_idr ASC")
    all_houses = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('properties.html', all_houses=all_houses)

@app.route('/properties-single/', defaults={'house_id': None})
@app.route('/properties-single/<house_id>')
def property_single_page(house_id):
    house_data = None
    if house_id:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM houses WHERE id = ?", (house_id,))
        house_data = cursor.fetchone()
        conn.close()
        if house_data is not None:
            house_data = dict(house_data) 

    if house_data is None:
        # Data rumah statis (Mulyorejo) jika tidak ada ID atau ID tidak ditemukan
        house_data = {
            "id": "mulyorejo_static", 
            "nama": "Mulyorejo",
            "harga_idr": 750000000,
            "lokasi": "Jalan Mulyosari Utara, Surabaya, ID",
            "kamar_tidur": 3,
            "kamar_mandi": 2,
            "luas_bangunan_m2": 1250,
            "luas_tanah_m2": 1300,
            "tahun_pembuatan": 2019,
            "image_url": "images/work-1.jpg" 
        }
    
    return render_template('properties-single.html', house=house_data)

@app.route('/hasil-simulasi')
def hasil_simulasi_page():
    return render_template('hasil-simulasi.html')

@app.route('/predict', methods=['POST'])
def predict():
    if request.method == 'POST':
        try:
            processed_results = preprocess_kpr_input(request.form) # Use the specific KPR preprocessing function
            processed_data_for_model = processed_results["model_input"]
            display_data_for_frontend = processed_results["display_data"]

            prediction_proba = model.predict_proba(processed_data_for_model)[0]
            prediction_label = model.predict(processed_data_for_model)[0] 
            
            # Assuming 1 is 'approved' and 0 is 'rejected' from the KPR model based on previous discussions.
            # If your model's output convention is 0 for Approved, 1 for Rejected, then:
            # result_text = "Disetujui" if prediction_label == 0 else "Ditolak"
            # probability_approved = prediction_proba[0] 
            # probability_refused = prediction_proba[1] 
            
            # Based on common ML classification for 'Approved' being 1 (positive class)
            # If your model returns 1 for approved, 0 for rejected
            result_text = "Disetujui" if prediction_label == 1 else "Ditolak"
            probability_approved = prediction_proba[1] # Probability of the positive class (1)
            probability_refused = prediction_proba[0] # Probability of the negative class (0)


            # KPR Simulation Calculation (in IDR)
            kpr_simulation_results = {}
            recommended_houses = []

            if prediction_label == 1: # Only calculate KPR and recommend if approved
                principal_loan_amount_idr = display_data_for_frontend.get('JumlahPinjamanDiajukan', 0)
                loan_term_months = display_data_for_frontend.get('Loan_Amount_Term', 0)

                kpr_simulation_results = calculate_kpr_simulation_details(
                    principal_loan_amount_idr, 
                    loan_term_months, 
                    KPR_ANNUAL_INTEREST_RATE
                )
                
                # For recommendations, use the IDR values
                user_income_total_idr_for_rec = display_data_for_frontend['ApplicantIncome'] + display_data_for_frontend['CoapplicantIncome']
                property_area_pref_for_rec = display_data_for_frontend.get('Property_Area', None)
                
                recommended_houses = get_recommended_houses(
                    user_income_total_idr_for_rec,
                    display_data_for_frontend['JumlahPinjamanDiajukan'],
                    property_area_pref_for_rec
                )

            response_message = {
                "status": "success",
                "prediction": result_text,
                "probability_approved": f"{probability_approved:.2f}",
                "probability_refused": f"{probability_refused:.2f}",
                "loan_details": display_data_for_frontend, # Send the IDR display data
                "kpr_simulation": kpr_simulation_results, # Send KPR simulation results
                "recommended_houses": recommended_houses 
            }
            return jsonify(response_message)

        except Exception as e:
            print(f"Error during prediction: {e}")
            import traceback
            traceback.print_exc() 
            return jsonify({"status": "error", "message": f"Terjadi kesalahan: {e}"}), 500

@app.route('/download-simulasi-kpr')
def download_simulasi_kpr():
    try:
        # Get parameters from query string
        principal_loan = float(request.args.get('principal'))
        loan_term = int(request.args.get('term'))
        interest_rate = float(request.args.get('rate', KPR_ANNUAL_INTEREST_RATE))
        
        # Calculate installments
        installments = calculate_monthly_installments(principal_loan, loan_term, interest_rate)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Simulasi KPR"

        # Styling
        header_fill = PatternFill(start_color="D8C75B", end_color="D8C75B", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title and info
        ws['A1'] = "SIMULASI KPR"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = "Jumlah Pinjaman:"
        ws['B3'] = f"Rp {principal_loan:,.2f}"
        ws['A4'] = "Jangka Waktu:"
        ws['B4'] = f"{loan_term} Bulan"
        ws['A5'] = "Suku Bunga:"
        ws['B5'] = f"{interest_rate}% per tahun"

        # Headers - Row 7
        headers = ['Bulan ke-', 'Cicilan Bulanan', 'Pembayaran Pokok', 'Pembayaran Bunga', 'Sisa Pinjaman']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row, inst in enumerate(installments, 8):
            ws.cell(row=row, column=1, value=inst['Bulan_ke']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=inst['Cicilan_Bulanan']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=inst['Pembayaran_Pokok']).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=inst['Pembayaran_Bunga']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=inst['Sisa_Pinjaman']).number_format = '#,##0.00'
            
            # Add borders to all cells in the row
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulasi_kpr_{timestamp}.xlsx"

        return Response(
            excel_file.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    app.run(debug=True)